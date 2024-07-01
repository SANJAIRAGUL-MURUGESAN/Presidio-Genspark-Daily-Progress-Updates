import csv
import pandas as pd
from datetime import datetime
from validate_email import validate_email
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class Employee:
    def __init__(self, name, dob, phone, email):
        self.name = name
        self.dob = dob
        self.phone = phone
        self.email = email
        self.age = self.calculate_age()

    def calculate_age(self):
        today = datetime.today()
        age = today.year - self.dob.year - ((today.month, today.day) < (self.dob.month, self.dob.day))
        return age
    
class EmployeeManager:
    def __init__(self):
        self.employees = []
        self.load_employees_from_file()
    
    def load_employees_from_file(self):
        try:
            with open('employees.txt', 'r') as f:
                for line in f:
                    if line.strip():  # Ignore empty lines
                        data = line.strip().split(', ')
                        name = data[0].split(': ')[1]
                        dob = datetime.strptime(data[1].split(': ')[1], '%Y-%m-%d')
                        phone = data[2].split(': ')[1]
                        email = data[3].split(': ')[1]
                        self.employees.append(Employee(name, dob, phone, email))
        except FileNotFoundError:
            print("No existing employee data found in employees.txt.")
        except Exception as e:
            print(f"Error loading data from employees.txt: {e}")
    
    def save_to_text(self):
        try:
            with open('employees.txt', 'a') as f:
                for employee in self.employees:
                    f.write(f"Name: {employee.name}, DOB: {employee.dob.strftime('%Y-%m-%d')}, "
                            f"Phone: {employee.phone}, Email: {employee.email}, Age: {employee.age}\n")
            print("Data appended to employees.txt")
        except Exception as e:
            print(f"Error saving data to employees.txt: {e}")

    def save_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(['Name', 'DOB', 'Phone', 'Email', 'Age'])
            for employee in self.employees:
                ws.append([employee.name, employee.dob, employee.phone, employee.email, employee.age])
            # Adjust column width for DOB column
            dob_column = ws['B']
            max_length = 0
            for cell in dob_column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions['B'].width = max_length + 2
            wb.save('employees.xlsx')
            print("Data saved to employees.xlsx")
        except Exception as e:
            print(f"Error saving data to employees.xlsx: {e}")

    def save_to_pdf(self):
        try:
            c = canvas.Canvas("employees.pdf", pagesize=letter)
            width, height = letter

            # Set the font and size
            c.setFont("Helvetica", 12)
            
            # Write header
            c.drawString(50, height - 50, "Employee Details")
            c.drawString(50, height - 70, "-" * 50)
            
            # Write employee details
            text_y = height - 100
            for employee in self.employees:
                c.drawString(50, text_y, f"Name: {employee.name}")
                c.drawString(250, text_y, f"DOB: {employee.dob.strftime('%Y-%m-%d')}")
                c.drawString(450, text_y, f"Phone: {employee.phone}")
                c.drawString(50, text_y - 20, f"Email: {employee.email}")
                c.drawString(250, text_y - 20, f"Age: {employee.age}")
                text_y -= 40

            c.save()
            print("Data saved to employees.pdf")
        except Exception as e:
            print(f"Error saving data to employees.pdf: {e}")

    def add_employee(self, name, dob_str, phone, email):
        try:
            dob = datetime.strptime(dob_str, '%Y-%m-%d')
        except ValueError:
            raise ValueError("Invalid date format. Please enter in YYYY-MM-DD format.")
        
        if len(phone) != 10 or not phone.isdigit():
            raise ValueError("Invalid phone number format. Please enter 10 digits.")
        
        if not validate_email(email):
            raise ValueError("Invalid email format. Please enter a valid email address.")
        
        employee = Employee(name, dob, phone, email)
        self.employees.append(employee)
        print(f"Employee '{name}' added successfully.")
    
def main():
    emp_manager = EmployeeManager()
    
    while True:
        print("\n--- Employee Details Input ---")
        try:
            name = input("Enter Your Name: ")
            dob_str = input("Enter Date of Birth (YYYY-MM-DD): ")
            phone = input("Enter Phone Number (10 digits): ")
            email = input("Enter Email Address: ")

            emp_manager.add_employee(name, dob_str, phone, email)


            another = input("Do you want to add another employee? (yes/no): ").lower()
            if another != 'yes':
                break

        except ValueError as ve:
            print(f"Error: {ve}. Please try again.")

    print("\n--- Save Data ---")
    print("Choose file format to save:")
    print("1. Text")
    print("2. Excel")
    print("3. PDF")

    choice = input("Enter your choice (1/2/3): ")

    try:
        if choice == '1':
            emp_manager.save_to_text()
        elif choice == '2':
            emp_manager.save_to_excel()
        elif choice == '3':
            emp_manager.save_to_pdf()
        else:
            print("Invalid file format. Data not saved.")
    except Exception as e:
        print(f"Error: {e}. Data not saved.")


if __name__ == "__main__":
    main()